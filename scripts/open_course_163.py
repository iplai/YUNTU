# coding=utf-8
import re
import os
import requests
import json
import time
from openpyxl import Workbook, load_workbook
from scripts.yuntu_doc import *

'''
网易公开课有7858个课程28303个视频
https://c.open.163.com/search/search.htm?query=#/search/all
'''


# 得到7868个course
def get_courses():
    courses = []
    for i in range(2):
        r = requests.get(
            'http://c.open.163.com/search/school.do?callback=&school=&pageNum=' + str(i + 1) + '&pSize=4000')
        courses += json.loads(r.text.strip()[1:-1], encoding='utf-8')['result']['dtos']
    import codecs
    json.dump(courses, codecs.open('courses.json', 'w', 'utf-8'), indent=4)


def get_courses_from_dwr():
    dwr_file = open('course.dwr' + '.txt', 'a')
    for i in range(4):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchCourse',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第一页
            'c0-param2': 'number:2000',  # 3000条每页，超过5000取不到数据
            'batchId': '1493367775649'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchCourse.dwr', data=data)
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


def get_special_from_dwr():
    dwr_file = open('special.dwr' + '.txt', 'a')
    for i in range(1):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchSpecial',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第一页
            'c0-param2': 'number:500',  # 500条每页，超过5000取不到数据
            'batchId': '1493367775649'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchSpecial.dwr', data=data)
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


def get_videos_from_dwr():
    """
    网易的dwr的规则是offset和limit都不能超过5000所以最多只能取9999条数据
    :return:
    """
    dwr_file = open('videos_dwr' + '.txt', 'a')
    for i in range(2):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '${scriptSessionId}190',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchVideo',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第i+1页
            'c0-param2': 'number:4999',
            'batchId': '1493470611201'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchVideo.dwr', data=data)
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


# 处理后（courseId唯一）有5945个课程，存入mongodb中
def process_courses_dwr():
    courses = []
    dwr_file = open('courses_dwr' + '.txt')
    for line in dwr_file:
        data = {}
        bigPicUrl = re.search(r'bigPicUrl="(.*?)"', line)
        if bigPicUrl:
            data['bigPicUrl'] = bigPicUrl.group(1).strip()
        category = re.search(r'category="(.*?)"', line)
        if category:
            data['category'] = category.group(1).strip().decode('unicode-escape').encode('utf-8')
        courseId = re.search(r'courseId="(.*?)"', line)
        if courseId:
            data['courseId'] = courseId.group(1).strip()
        courseUrl = re.search(r'courseUrl="(.*?)"', line)
        if courseUrl:
            data['url'] = courseUrl.group(1).strip()
        description = re.search(r'description="(.*?)"', line)
        if description:
            data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
        instructor = re.search(r'instructor="(.*?)"', line)
        if instructor:
            data['instructor'] = instructor.group(1).strip().decode('unicode-escape').encode('utf-8')
        movieCount = re.search(r'movieCount=(\d+)', line)
        if movieCount:
            data['totalMovieCount'] = int(movieCount.group(1))
        school = re.search(r'school="(.*?)"', line)
        if school:
            data['school'] = school.group(1).strip().decode('unicode-escape').encode('utf-8')
        subject = re.search(r'subject="(.*?)"', line)
        if subject:
            data['subject'] = subject.group(1).strip().decode('unicode-escape').encode('utf-8')
        tags = re.search(r'tags="(.*?)"', line)
        if tags:
            data['tags'] = tags.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        if title:
            data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        courses.append(data)
    return courses


def process_special_dwr():
    courses = []
    dwr_file = open('courses_dwr' + '.txt')
    for line in dwr_file:
        data = {}
        bigPicUrl = re.search(r'bigPicUrl="(.*?)"', line)
        if bigPicUrl:
            data['bigPicUrl'] = bigPicUrl.group(1).strip()
        category = re.search(r'category="(.*?)"', line)
        if category:
            data['category'] = category.group(1).strip().decode('unicode-escape').encode('utf-8')
        courseId = re.search(r'courseId="(.*?)"', line)
        if courseId:
            data['courseId'] = courseId.group(1).strip()
        courseUrl = re.search(r'courseUrl="(.*?)"', line)
        if courseUrl:
            data['url'] = courseUrl.group(1).strip()
        description = re.search(r'description="(.*?)"', line)
        if description:
            data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
        instructor = re.search(r'instructor="(.*?)"', line)
        if instructor:
            data['instructor'] = instructor.group(1).strip().decode('unicode-escape').encode('utf-8')
        movieCount = re.search(r'movieCount=(\d+)', line)
        if movieCount:
            data['totalMovieCount'] = int(movieCount.group(1))
        school = re.search(r'school="(.*?)"', line)
        if school:
            data['school'] = school.group(1).strip().decode('unicode-escape').encode('utf-8')
        subject = re.search(r'subject="(.*?)"', line)
        if subject:
            data['subject'] = subject.group(1).strip().decode('unicode-escape').encode('utf-8')
        tags = re.search(r'tags="(.*?)"', line)
        if tags:
            data['tags'] = tags.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        if title:
            data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        courses.append(data)
    return courses


def process_videos_dwr():
    videos = []
    dwr_file = open('videos_dwr' + '.txt')
    for line in dwr_file:
        data = {}
        picUrl = re.search(r'imgurl=(.*?)_180x100x1x95.jpg', line)
        data['bigPicUrl'] = picUrl.group(1).strip()
        url = re.search(r'url="(.*?)"', line)
        data['url'] = url.group(1).strip()
        tem = re.search(r'^(.*?)\.html', data['url'].split('/')[-1]).group(1)
        data['courseId'] = tem.split('_')[0]
        data['videoId'] = tem.split('_')[1]
        description = re.search(r'description="(.*?)"', line)
        data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        videos.append(data)
    return videos


def insert_courses():
    i = 0
    courses = json.load(open('courses.json'), encoding='utf-8')
    from mongoengine.errors import NotUniqueError
    for course in courses:
        for k, v in course.items():
            if type(v) in [str, unicode]:
                if len(v) == 0:
                    del course[k]
                else:
                    course[k] = v.strip()
        if 'courseUrl' in course:
            try:
                i += 1
                print i, course['title']
                Course(**course).save()
            except NotUniqueError, e:
                print e


def insert_videos():
    i = 0
    for course in Course.objects:
        for url in course.videoUrls:
            try:
                data = {
                    'url': url,
                    'courseId': course.courseId,
                }
                Video(**data).save()
                i += 1
                print i
            except NotUniqueError, e:
                print e


def process_data(cate):
    f = open(cate + '.txt')
    xls_file = 'courses.xlsx'
    if os.path.isfile(xls_file):
        wb = load_workbook(filename=xls_file)
    else:
        wb = Workbook()
    ws = wb.create_sheet(title=cate.decode('utf-8'))
    ws.append(
        ['title', 'category', 'cover', 'courseId', 'courseType', 'courseUrl', 'description', 'instructor', 'movieCount',
         'school', 'startTime', 'subject', 'tags'])
    for l in f:
        a = re.sub(r';s\d+\.', "|**|", l).replace('null', 'None').replace('{##', '').replace('##}', '')
        a = re.sub(r's\d+\.', "", a, 1)
        for i in a.split('|**|'):
            # print i
            exec i
        description = description.decode('unicode-escape').encode('utf-8') if description else None
        instructor = instructor.decode('unicode-escape').encode('utf-8') if instructor else None
        school = school.decode('unicode-escape').encode('utf-8') if school else None
        subject = subject.decode('unicode-escape').encode('utf-8') if subject else None
        tags = tags.decode('unicode-escape').encode('utf-8') if tags else None
        category = category.decode('unicode-escape').encode('utf-8') if category else None
        title = title.decode('unicode-escape').encode('utf-8') if title else None
        cover = bigPicUrl
        state = 30
        type = 1
        ws.append(
            [title, category, cover, courseId, courseType, courseUrl, description, instructor, movieCount, school,
             startTime, subject, tags])
    wb.save(xls_file)


def export_data():
    xls_file = '2.xlsx'
    wb = load_workbook(filename=xls_file)
    album_sheet = wb.get_sheet_by_name(u'专辑')
    media_sheet = wb.get_sheet_by_name(u'视频')
    for course in Course.objects(crawled=True):
        print course.title
        label1 = 230000
        label2 = 230100 if 'TED' == course.school else 230200
        if course.tags and 'TED' in course.tags:
            label2 = 230100
        data = [
            course.courseId, course.title, course.bigPicUrl, label1, label2, course.tags, course.description, '公开课',
            '57cd06e877c8a33013933976', '57e10a5ea770286d5ebd4878', course.instructor, '视频', '网易公开课'
        ]
        album_sheet.append(data)
        for item in course.videoUrls:
            try:
                video = Video.objects.get(url=item)
                print video.title
                course = Course.objects.get(courseId=video.courseId)
                url = video.videoUrl
                if url:
                    if '-list.m3u8' in url:
                        url = url.replace('-list.m3u8', '.flv')
                    else:
                        url = url.replace('.m3u8', '.mp4')
                if not video.bigPicUrl:
                    cover = None
                else:
                    if 'oimagec2.ydstatic.com' in video.bigPicUrl:
                        cover = re.search(r'url=(.*\.jpg)', video.bigPicUrl).group(1)
                    elif 'imgsize.ph.126.net' in video.bigPicUrl:
                        cover = re.search(r'url=(.*\.jpg)_', video.bigPicUrl).group(1)
                    else:
                        cover = video.bigPicUrl
                media_sheet.append([
                    course.courseId, video.title, cover, course.tags, video.description, u'公开课',
                    '57cd06e877c8a33013933976', '57e10a5ea770286d5ebd4878', course.instructor, u'视频', u'网易公开课', url
                ])
            except:
                pass
    wb.save(xls_file)


# 7148
def insert_media_album():
    count = 0
    for course in Course.objects():
        tags = course.tags
        seps = [',', ';', u'，', u'；', u'、']
        if not tags:
            pass
        else:
            for i in seps:
                if tags.endswith(i):
                    tags = tags[:-1]
            for i in seps:
                if i in tags:
                    tags = [tag for tag in tags.split(i) if tag != '']
                    break
            if type(tags) != list:
                tags = [tags]
        data = dict(
            title=course.title,
            cover=course.cover,
            timeline=long(round(time.time() * 1000)),  # 当前时间的13位时间戳
            introduction=course.description,
            tags=tags,
            source=u'网易公开课',
            uploadUsername=u'oc',
            uploadUid='57cd06e877c8a33013933976',
            lid='57cd06c677c8a33013933975',
            mediaCount=len(course.videoUrls),
            state=30, comment=0,
            favorite=0,
            type=1,
            hot=0,
            open=1,
            channel=0,
            view=0,
            artist=course.instructor,
            label1=230000,
            label2=230100 if 'TED' == course.school else 230200
        )
        media_album = MediaAlbum(**data)
        media_album.save()
        course.update(aid=str(media_album['id']))
        print media_album.tags
        count += 1
        print count, str(media_album['id']), 'mongodb saved!', data['title']

        es.index('media-album-index',
                 'media-album',
                 dict(data, id=str(media_album['id'])),
                 id=str(media_album['id']))
        print count, str(media_album['id']), 'es saved!', data['title']


def insert_media():
    count = 0
    total = Video.objects(crawled=True).count()
    for video in Video.objects(crawled=True):
        course = Course.objects.get(courseId=video.courseId)
        media_album = MediaAlbum.objects.get(id=course.aid)
        if not video.bigPicUrl:
            cover = None
        else:
            if 'oimagec2.ydstatic.com' in video.bigPicUrl:
                cover = re.search(r'url=(.*\.jpg)', video.bigPicUrl).group(1)
            elif 'imgsize.ph.126.net' in video.bigPicUrl:
                cover = re.search(r'url=(.*\.jpg)_', video.bigPicUrl).group(1)
            else:
                cover = video.bigPicUrl
        url = video.videoUrl
        if url:
            if '-list.m3u8' in url:
                url = url.replace('-list.m3u8', '.flv')
            else:
                url = url.replace('.m3u8', '.mp4')
        data = dict(
            title=video.title,
            url=url,
            timeline=long(round(time.time() * 1000)),  # 当前时间的13位时间戳
            tags=media_album.tags,
            lid='57cd06c677c8a33013933975',
            cover=cover,
            state=30,
            type=1,
            open=1,
            aid=course.aid,
        )
        media = Media(**data)
        media.save()
        video.update(mid=str(media.id))
        count += 1
        print str(count) + '/' + str(total)


def modify_video_url():
    for media in Media.objects(timeline__gt=1493696034175):
        url = media.url
        if url:
            if '-list.m3u8' in url:
                url = url.replace('-list.m3u8', '.flv')
            else:
                url = url.replace('.m3u8', '.mp4')
            print url
            media.update(url=url)


def search_video_by_course_id(id, keyword, limit):
    data = {
        'callCount': 1,
        'scriptSessionId': '${scriptSessionId}190',
        'httpSessionId': '',
        'c0-scriptName': 'OpenSearchBean',
        'c0-methodName': 'searchVideo',
        'c0-id': 0,
        'c0-param0': 'string:' + keyword,
        'c0-param1': 'number:1',  # 第一页
        'c0-param2': 'number:' + str(limit),  # 3000条每页，超过3000取不到数据
        'batchId': '1493470611201'
    }
    r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchVideo.dwr', data=data)
    items = '\n'.join(r.text.split('\n')[4:-2])
    if '' != items:
        for item in items.split('\n'):
            videoUrl = re.search(r'url="(.*?)"', item).group(1).strip()
            if id in videoUrl:
                return videoUrl
    return None


def delete_media_album():
    MediaAlbum.objects(source='网易公开课').delete()
    a = es.search(index="media-album-index", body={'size': 8000, 'query': {'match': {'source': '网易公开课'}}})
    for i in a['hits']['hits']:
        es.delete('media-album-index', 'media-album', i['_id'])
        print i['_id']


if __name__ == "__main__":
    cates = ['TED', 'BBC', '可汗学院', '国际名校公开课', '中国大学视频公开课', '国立台湾大学公开课']
    for cate in cates:
        print cate
        # get_data(cate)
        # process_data(cate)
